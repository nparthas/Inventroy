import SQLinventory as sq
import tkinter as tk
from tkinter import messagebox
# from tkinter import ttk
from sqlite3 import OperationalError
import openpyxl as xl
import time

import logging

logging.basicConfig(level=logging.DEBUG, format=' %(asctime)s - %(levelname)s - %(message)s')
# logging.disable(logging.CRITICAL)
# logging.disable(logging.DEBUG)

logging.debug('Start of program')

# from functools import reduce


TITLE_FONT = ("Myriad Pro", 18, "bold")


def multi_function(*functions):  # for multiple functions inside one button
    def func(*args, **kwargs):
        values = None
        for function in functions:
            values = function(*args, **kwargs)
        return values

    return func


"""
def on_closing():
    if messagebox.askokcancel("Quit", "Do you want to quit?"):
        tk.destroy()
"""


class InventoryApp(tk.Tk):
    connection = sq.createDB('Z:\Inventory\InventoryGUI\inventory.db')

    try:
        work_book = xl.load_workbook('inventory.xlsx')
    except FileNotFoundError:
        work_book = xl.Workbook()

    def __init__(self, *args, **kwargs):
        tk.Tk.__init__(self, *args, **kwargs)

        # the container is where we'll stack a bunch of frames
        # on top of each other, then the one we want visible
        # will be raised above the others
        container = tk.Frame(self)
        container.pack(side="top", fill="both", expand=True)
        container.grid_rowconfigure(0, weight=1)
        container.grid_columnconfigure(0, weight=1)

        if not self.connection:
            messagebox.showwarning('ERROR', 'Unable to connect to database')

        if self.connection:
            self.protocol("WM_DELETE_WINDOW", self.on_closing)

        self.title('Amphenol Commercial I/O')
        self.iconbitmap('Amphenol-Icon.ico')

        self.frames = {}
        for F in (MainPage, MakeTables, TableList, ViewTables, AddValues):
            page_name = F.__name__
            frame = F(parent=container, controller=self)
            self.frames[page_name] = frame

            # put all of the pages in the same location;
            # the one on the top of the stacking order
            # will be the one that is visible.
            frame.grid(row=0, column=0, sticky="nsew")
            frame.columnconfigure(0, weight=1)
            frame.rowconfigure(0, weight=1)
            # frame.configure(background='#ebf1f8') FIX LATER

        self.show_frame("MainPage")

    def show_frame(self, page_name):
        # Show a frame for the given page name
        frame = self.frames[page_name]
        frame.tkraise()

    def on_closing(self):
        # if messagebox.askokcancel("Quit", "Do you want to quit?"):
        InventoryApp.connection.commit()
        InventoryApp.connection.close()
        InventoryApp.work_book.save('Inventory.xlsx')

        self.destroy()


class MainPage(tk.Frame):
    def __init__(self, parent, controller):
        tk.Frame.__init__(self, parent)
        self.controller = controller
        # self.configure(background='#0062a6')

        title = tk.Label(self, text="Amphenol Sample Inventory Database", font=TITLE_FONT, fg='#0062a6')
        title.pack(side="top", fill="x", pady=(10, 20))

        make_tables = tk.Button(self, text='Make Tables', width=20,
                                command=lambda: controller.show_frame("MakeTables"))
        table_list = tk.Button(self, text="View Table List", width=30,
                               command=lambda: controller.show_frame("TableList"))
        view_tables = tk.Button(self, text="View Table Contents", width=30,
                                command=lambda: controller.show_frame("ViewTables"))
        add_values = tk.Button(self, text="Add Values", width=30,
                               command=lambda: controller.show_frame('AddValues'))

        if InventoryApp.connection:
            close = tk.Button(self, text="Exit", width=15,
                              command=lambda: InventoryApp.on_closing(app))
        else:
            close = tk.Button(self, text='Exit', width=15, command=lambda: multi_function(
                InventoryApp.work_book.save('Inventory.xlsx'), exit()))

        make_tables.place(rely=1.0, relx=1.0, x=0, y=0, anchor=tk.SE)
        table_list.pack()
        view_tables.pack()
        add_values.pack()
        close.pack(side='bottom')


class MakeTables(tk.Frame):
    def __init__(self, parent, controller):
        tk.Frame.__init__(self, parent)
        self.controller = controller

        title = tk.Label(self, text="Initialize Tables", font=TITLE_FONT, fg='#0062A6')
        title.pack(anchor=tk.N, side=tk.TOP, fill=tk.X, pady=(5, 10))

        result = tk.Label(self, text="")
        result.pack(side=tk.TOP, pady=(0, 10))

        button_frame = tk.Frame(self)
        button_frame.pack(anchor=tk.CENTER, side=tk.TOP, fill=tk.X)

        table_list = tk.Button(button_frame, text="Initialize Table List", width=40,
                               command=lambda: result.configure(text=sq.createTableList(InventoryApp.connection)))

        connectors_table = tk.Button(button_frame, text='Initialize Connectors Table', width=40,
                                     command=lambda: result.configure(
                                         text=sq.createConnectorsTable(InventoryApp.connection)))

        connectors_history_table = tk.Button(button_frame, text="Initialize Connectors History Table", width=40,
                                             command=lambda: result.configure(
                                                 text=sq.createConnectorsHistoryTable(InventoryApp.connection)))

        sample_cases_table = tk.Button(button_frame, text="Initialize Sample Cases Table", width=40,
                                       command=lambda: result.configure(
                                           text=sq.createSampleCasesTable(InventoryApp.connection)))

        sample_cases_history_table = tk.Button(button_frame, text="Initialize Sample Cases History Table", width=40,
                                               command=lambda: result.configure(
                                                   text=sq.createSampleCasesHistoryTable(InventoryApp.connection)))
        changelog_table = tk.Button(button_frame, text='Initialize Changelog Table', width=40,
                                    command=lambda: result.configure(
                                        text=sq.create_changelog_table(InventoryApp.connection)))

        table_list.pack()
        connectors_table.pack()
        connectors_history_table.pack()
        sample_cases_table.pack()
        sample_cases_history_table.pack()
        changelog_table.pack()

        home = tk.Button(self, text="Go to the main page", command=lambda: controller.show_frame("MainPage"))
        home.pack(anchor=tk.S, side=tk.LEFT, pady=(15, 0))


class TableList(tk.Frame):
    def __init__(self, parent, controller):
        tk.Frame.__init__(self, parent)
        self.controller = controller

        try:
            sq.trim_TableList(InventoryApp.connection)

        except OperationalError:
            messagebox.showwarning('ERROR', 'Database is locked')

        title = tk.Label(self, text='List of All Tables: ', font=TITLE_FONT, fg='#0062A6')
        title.pack(anchor=tk.N, side=tk.TOP, fill=tk.X, pady=(10, 30))

        self.table_frame = tk.Frame(self)
        self.table_frame.pack(anchor=tk.N, side=tk.TOP)

        home = tk.Button(self, text="Go to the main page", command=lambda: controller.show_frame("MainPage"))
        home.pack(anchor=tk.S, side=tk.LEFT)

        refresh = tk.Button(self, text='Refresh list',
                            command=lambda: self.refresh_values(
                                InventoryApp.connection.execute('''SELECT TableListID as 'Table List ID',
                                                                Name,
                                                                DateModified as 'Date Modified'
                                                                FROM TableList''')))
        refresh.pack(anchor=tk.S, side=tk.RIGHT)

        self.refresh_values(
            InventoryApp.connection.execute('''SELECT TableListID as 'Table List ID',
                                                                Name,
                                                                DateModified as 'Date Modified'
                                                                FROM TableList'''))

    def refresh_values(self, connection_object):
        for widget in self.table_frame.winfo_children():
            widget.destroy()

        if connection_object is not None:
            i = 0
            j = 0
            for item in connection_object.description:
                spot = tk.Label(self.table_frame, text=item[0], anchor=tk.W, bg="#B8E1EB")
                spot.grid(row=i, column=j, sticky=tk.N + tk.E + tk.S + tk.W)
                j += 1
            i += 1

            for row in connection_object:
                j = 0
                for value in row:
                    spot = tk.Label(self.table_frame, text=value, anchor=tk.W)
                    spot.grid(row=i, column=j, sticky=tk.N + tk.E + tk.S + tk.W)
                    # if i % 2 == 1:
                    #    spot.configure(bg='#b8e1eb')
                    j += 1
                i += 1
            for k in range(3):
                padding_row = tk.Label(self.table_frame, text="", padx=75)
                padding_row.grid(row=i, column=k, sticky=tk.W + tk.E)


class ViewTables(tk.Frame):
    def __init__(self, parent, controller):

        def get_entry(*args):
            if args is not None:
                self.entry_temp_list = []
                for item in args:
                    self.entry_temp_list.append(item.get())
                    item.delete(0, tk.END)
                refresh_entry()
            return self.entry_temp_list
            # could remove the temp list if found to be useless

        def refresh_entry():
            variable_lookup_column.insert(0, 'Input specifier type')
            variable_lookup_criteria.insert(0, 'Input specifier value')
            variable_lookup_table.insert(0, 'Input table')

        def on_frame_configure(event):
            info_grid_canvas.configure(scrollregion=info_grid_canvas.bbox("all"))

        def create_sheet():
            self.work_sheet = InventoryApp.work_book.create_sheet()
            self.work_sheet.title = 'Inventory{0} '.format(time.strftime('%d-%m-%y'))

        def print_to_excel():
            create_sheet()

            try:
                num_cols = self.label_information.cget('text').split()[1]
                self.label_information.destroy()

                current_row = 1
                current_col = 1

                for widget in self.info_grid_frame.winfo_children():
                    cell = self.work_sheet.cell(row=current_row, column=current_col)
                    cell.value = widget['text']

                    current_col += 1

                    if current_col > int(num_cols):
                        current_col = 1
                        current_row += 1

                InventoryApp.work_book.save('Inventory.xlsx')
                excel_result.configure(text='Results printed')

            except Exception:
                excel_result.configure(text='Error: results not printed')

        tk.Frame.__init__(self, parent)
        self.controller = controller
        self.label_information = None  # For initialization
        self.work_sheet = None  # For initialization

        title_frame = tk.Frame(self)
        title_frame.pack(anchor=tk.N, side=tk.TOP, fill=tk.X, pady=(5, 20))

        home_frame = tk.Frame(self)
        home_frame.pack(anchor=tk.S, side=tk.BOTTOM, fill=tk.X)

        view_tables_frame = tk.Frame(self)
        view_tables_frame.pack(anchor=tk.N, side=tk.TOP, fill=tk.X)

        for i in range(4):
            view_tables_frame.columnconfigure(i, weight=2)

        entry_frame = tk.Frame(self)
        entry_frame.pack(anchor=tk.N, side=tk.TOP, pady=15, fill=tk.X)

        for i in range(4):
            entry_frame.columnconfigure(i, weight=2)

        info_grid_canvas = tk.Canvas(self, borderwidth=0)

        self.info_grid_frame = tk.Frame(info_grid_canvas)
        self.info_grid_frame.pack(anchor=tk.N, side=tk.TOP, fill=tk.BOTH)

        info_grid_scrollbar = tk.Scrollbar(self, orient='vertical', command=info_grid_canvas.yview)

        info_grid_canvas.configure(yscrollcommand=info_grid_scrollbar.set)

        info_grid_scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        info_grid_canvas.pack(side=tk.TOP, fill=tk.BOTH, expand=True)
        info_grid_canvas.create_window((4, 4), window=self.info_grid_frame, anchor=tk.NW, tags='info_grid_frame')

        self.info_grid_frame.bind("<Configure>", on_frame_configure)

        home = tk.Button(home_frame, text="Go to the main page", command=lambda: controller.show_frame("MainPage"))
        home.pack(anchor=tk.S, side=tk.LEFT)

        title = tk.Label(title_frame, text='View Table Contents', font=TITLE_FONT, fg='#0062A6')
        title.pack()

        view_connectors = tk.Button(view_tables_frame, text='View Connectors',
                                    command=lambda: self.refresh_values(
                                        InventoryApp.connection.execute('SELECT * FROM Connectors')))
        view_connectors.grid(column=0, row=0, sticky=tk.W + tk.E)

        view_connectors_history = tk.Button(view_tables_frame, text='View Connectors History',
                                            command=lambda: self.refresh_values(
                                                InventoryApp.connection.execute(
                                                    'SELECT * FROM ConnectorsHistory')))
        view_connectors_history.grid(column=2, row=0, sticky=tk.W + tk.E)

        view_sample_cases = tk.Button(view_tables_frame, text='View Sample Cases',
                                      command=lambda: self.refresh_values(
                                          InventoryApp.connection.execute(
                                              'SELECT * FROM SampleCases')))
        view_sample_cases.grid(column=1, row=0, sticky=tk.W + tk.E)

        view_sample_cases_history = tk.Button(view_tables_frame, text='View Sample Cases History',
                                              command=lambda: self.refresh_values(
                                                  InventoryApp.connection.execute(
                                                      'SELECT * FROM SampleCasesHistory')))
        view_sample_cases_history.grid(column=3, row=0, sticky=tk.W + tk.E)

        variable_lookup_text = tk.Label(entry_frame, text='Variable Lookup')
        variable_lookup_text.grid(column=0, row=0, sticky=tk.W + tk.E, padx=(0, 15))

        variable_lookup_column_label = tk.Label(entry_frame, text="Type")
        variable_lookup_column_label.grid(column=1, row=0, sticky=tk.W)

        variable_lookup_criteria_label = tk.Label(entry_frame, text="Criteria")
        variable_lookup_criteria_label.grid(column=2, row=0, sticky=tk.W)

        variable_lookup_table_label = tk.Label(entry_frame, text="Table")
        variable_lookup_table_label.grid(column=3, row=0, sticky=tk.W)

        variable_lookup_column = tk.Entry(entry_frame, width=30)
        variable_lookup_column.insert(0, "Input specifier type")
        variable_lookup_column.grid(column=1, row=1, sticky=tk.W)

        variable_lookup_criteria = tk.Entry(entry_frame, width=30)
        variable_lookup_criteria.insert(0, 'Input specifier value')
        variable_lookup_criteria.grid(column=2, row=1, sticky=tk.W)

        variable_lookup_table = tk.Entry(entry_frame, width=30)
        variable_lookup_table.insert(0, 'Input table')
        variable_lookup_table.grid(column=3, row=1, sticky=tk.W)

        enter_button = tk.Button(entry_frame, text='Enter',
                                 command=lambda: self.refresh_values(
                                     sq.query_table(InventoryApp.connection, get_entry(
                                         variable_lookup_column,
                                         variable_lookup_criteria,
                                         variable_lookup_table))))
        enter_button.bind("<Return>", (lambda event: self.refresh_values(
            sq.query_table(InventoryApp.connection, get_entry(
                variable_lookup_column,
                variable_lookup_criteria,
                variable_lookup_table)))))
        enter_button.grid(column=2, row=2, sticky=tk.W + tk.E)

        erase_button = tk.Button(entry_frame, text='Erase', command=lambda: self.refresh_values(""))
        erase_button.bind("<Return>", lambda event: self.refresh_values(""))

        erase_button.grid(column=3, row=2, sticky=tk.W + tk.E)

        print_to_excel_button = tk.Button(entry_frame, text="Print to Excel", command=print_to_excel)
        print_to_excel_button.bind("<Return>", lambda event: print_to_excel())

        print_to_excel_button.grid(column=3, row=3, sticky=tk.W + tk.E)

        excel_result = tk.Label(entry_frame, text="")
        excel_result.grid(column=2, row=3, sticky=tk.W + tk.E)

    def refresh_values(self, connection_object):
        for widget in self.info_grid_frame.winfo_children():
            widget.destroy()

        if connection_object == 'Error (Check Capitalization)':
            spot = tk.Label(self.info_grid_frame, text="Error (Check Capitalization)")
            spot.pack(anchor=tk.N, side=tk.TOP)
            return None

        if connection_object != '':
            i = 0
            j = 0
            for item in connection_object.description:
                spot = tk.Label(self.info_grid_frame, text=item[0])
                spot.grid(row=i, column=j, sticky=tk.N + tk.E + tk.S + tk.W, padx=7)
                j += 1
            i += 1
            j = 0

            for row in connection_object:
                j = 0
                for value in row:
                    spot = tk.Label(self.info_grid_frame, text=value)
                    spot.grid(row=i, column=j, sticky=tk.N + tk.E + tk.S + tk.W)
                    if i % 2 == 1:
                        spot.configure(bg='#b8e1eb')
                    j += 1
                i += 1

            self.label_information = tk.Label(self.info_grid_frame, fg='#F0F0ED', text="{0} {1}".format(i, j))
            self.label_information.grid(row=i + 1, column=j + 1)


class AddValues(tk.Frame):
    def __init__(self, parent, controller):
        tk.Frame.__init__(self, parent)
        self.controller = controller

        title_frame = tk.Frame(self)
        title_frame.pack(anchor=tk.N, side=tk.TOP, fill=tk.X, pady=(5, 30))

        self.table_to_modify_frame = tk.Frame(self)
        self.table_to_modify_frame.pack(anchor=tk.N, side=tk.TOP)

        self.table_to_modify_label = None  # For initialization
        self.table_to_modify_entry = None  # For initialization
        self.table_to_modify_enter_button = None  # For initialization

        self.table_to_modify()

        self.entry_frame = tk.Frame(self)
        self.entry_frame.pack(anchor=tk.N, side=tk.TOP, pady=(5, 0))

        title = tk.Label(title_frame, text='Edit Table Contents', font=TITLE_FONT, fg='#0062A6')
        title.pack()

        home = tk.Button(self, text="Go to the main page", command=lambda: controller.show_frame("MainPage"))
        home.pack(anchor=tk.S, side=tk.LEFT)

    def table_enter_button(self):
        table_name = self.table_to_modify_entry.get()

        try:
            headers_query = InventoryApp.connection.execute("PRAGMA table_info({0})".format(table_name)).fetchall()
            headers = [x[1] for x in headers_query]

            for widget in self.table_to_modify_frame.winfo_children():
                widget.destroy()

            j = 0
            for header in headers:
                column = tk.Label(self.entry_frame, text=header, anchor=tk.W)
                column.grid(row=0, column=j, sticky=tk.N + tk.E + tk.S + tk.W)
                j += 1

            j = 0
            for _ in enumerate(headers):
                entry = tk.Entry(self.entry_frame, width=15)
                entry.grid(row=1, column=j, sticky=tk.N + tk.E + tk.S + tk.W)
                j += 1

            enter_button = tk.Button(self.entry_frame, text='Enter', width=20,
                                     command=lambda: multi_function(
                                         sq.fillTable(
                                             self.prepare_dict(), InventoryApp.connection, table_name),
                                         self.destroy_entry_frame(),
                                         self.table_to_modify()))
            enter_button.grid(row=1, column=j, sticky=tk.N + tk.E + tk.S + tk.W)

        except OperationalError:
            error_message = tk.Label(self.entry_frame, text='Incorrect Table Name')
            error_message.grid(row=0, column=0)

    def table_to_modify(self):

        self.table_to_modify_label = tk.Label(self.table_to_modify_frame, text='Insert table to edit')
        self.table_to_modify_label.pack(side=tk.LEFT, anchor=tk.W, padx=10)

        self.table_to_modify_entry = tk.Entry(self.table_to_modify_frame, width=30)
        self.table_to_modify_entry.pack(side=tk.LEFT, anchor=tk.W)

        self.table_to_modify_enter_button = tk.Button(self.table_to_modify_frame, text='Enter', width=20,
                                                      command=lambda: self.table_enter_button())
        self.table_to_modify_enter_button.bind("<Return>", lambda event: self.table_enter_button())
        self.table_to_modify_enter_button.pack(side=tk.LEFT, anchor=tk.W)

    def destroy_entry_frame(self):
        for widget in self.entry_frame.winfo_children():
            widget.destroy()

    def prepare_dict(self):
        keys = []
        values = []
        for widget in self.entry_frame.winfo_children():
            if widget.winfo_class() == 'Label':
                keys.append(widget['text'])

            if widget.winfo_class() == 'Entry':
                values.append(widget.get())

        logging.debug(keys)
        logging.debug(values)

        statement_dict = dict(zip(keys, values))

        logging.debug(dict((key, value) for key, value in statement_dict.items() if value is not ''))

        return dict((key, value) for key, value in statement_dict.items() if value is not '')

    """
    ACTION LIST:

    1 insert table to edit
    2 entry widget disappears
    3 entry widgets for corresponding table appear
    4 entry/cancel button appears
    5 table entry widgets disappear
    6 table is updated along with changelog
    7 insert table to edit widget appears again


    """

if __name__ == "__main__":
    app = InventoryApp()
    app.mainloop()
